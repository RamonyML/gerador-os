import sys
import json
import openpyxl
from datetime import datetime, date

path = sys.argv[1]
wb = openpyxl.load_workbook(path, data_only=True)

print("=== SHEETS ===")
for s in wb.sheetnames:
    print(f"  {s}")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n=== SHEET: {sheet_name} (max_row={ws.max_row}, max_col={ws.max_column}) ===")
    # Print first 5 rows
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= 5:
            break
        vals = [str(v) if v is not None else '' for v in row]
        print(f"  row {i+1}: {vals}")
    # Count non-empty rows
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in row):
            count += 1
    print(f"  Total data rows: {count}")
